#!/usr/bin/env python3
"""
创建 AMP HC 汇总 Sheet v4
- 添加 Lodz
- Juarez → Queretaro
- Farmington Hills → Novi
- Karlsbad → Garching
- Shanghai → Suzhou
"""

import pandas as pd
import glob
import os
import math
from openpyxl import load_workbook

# 文件目录
data_dir = '/Users/liwang/projects/CA_AMP_Workshop_Report'
os.chdir(data_dir)

# 查找文件
files = [f for f in glob.glob("*AMP*HC*.xlsx") if not f.startswith('~$')]
file_path = os.path.abspath(files[0])
print(f"📂 读取文件：{file_path}")

# 读取原始数据
df_2025 = pd.read_excel(file_path, sheet_name='2025 Jan')
df_2026 = pd.read_excel(file_path, sheet_name='2026 Feb')

# 清理数据
df_2025['Int/Ext'] = df_2025['Int/Ext'].fillna('Unknown')
df_2026['Int/Ext'] = df_2026['Int/Ext'].fillna('Unknown')
df_2025['Location'] = df_2025['Location'].fillna('Unknown')
df_2026['Location'] = df_2026['Location'].fillna('Unknown')

# 地点名称统一
df_2025.loc[df_2025['Location'] == 'Juarez', 'Location'] = 'Queretaro'
df_2026.loc[df_2026['Location'] == 'Juarez', 'Location'] = 'Queretaro'
df_2025.loc[df_2025['Location'] == 'Farmington Hills', 'Location'] = 'Novi'
df_2026.loc[df_2026['Location'] == 'Farmington Hills', 'Location'] = 'Novi'
df_2025.loc[df_2025['Location'] == 'Karlsbad', 'Location'] = 'Garching'
df_2026.loc[df_2026['Location'] == 'Karlsbad', 'Location'] = 'Garching'
df_2025.loc[df_2025['Location'] == 'Shanghai', 'Location'] = 'Suzhou'
df_2026.loc[df_2026['Location'] == 'Shanghai', 'Location'] = 'Suzhou'
# Location=0 改为 Lodz
df_2025.loc[df_2025['Location'] == 0, 'Location'] = 'Lodz'
df_2026.loc[df_2026['Location'] == 0, 'Location'] = 'Lodz'

print("📍 已将 Juarez 归为 Queretaro")
print("📍 已将 Farmington Hills 改为 Novi")
print("📍 已将 Karlsbad 归为 Garching")
print("📍 已将 Shanghai 归为 Suzhou")

# 汇总 2025 Jan
summary_2025 = df_2025.groupby(['Location', 'Int/Ext'])['HC'].sum().unstack(fill_value=0)
if 'Internal' not in summary_2025.columns:
    summary_2025['Internal'] = 0
if 'External' not in summary_2025.columns:
    summary_2025['External'] = 0
summary_2025['Total_2025'] = summary_2025['Internal'] + summary_2025['External']

# 汇总 2026 Feb
summary_2026 = df_2026.groupby(['Location', 'Int/Ext'])['HC'].sum().unstack(fill_value=0)
if 'Internal' not in summary_2026.columns:
    summary_2026['Internal'] = 0
if 'External' not in summary_2026.columns:
    summary_2026['External'] = 0
summary_2026['Total_2026'] = summary_2026['Internal'] + summary_2026['External']

# 合并两年数据
all_locations = sorted(set(str(x) for x in summary_2025.index) | set(str(x) for x in summary_2026.index))

# 构建汇总 DataFrame
summary_data = []
for loc in all_locations:
    # 忽略 NaN 和空值，不统计在任何数据里
    if loc == 'Unknown' or loc == 'nan' or loc == '0' or (isinstance(loc, float) and math.isnan(loc)):
        continue
    if pd.isna(loc):
        continue
    
    int_2025 = int(summary_2025.loc[loc, 'Internal']) if loc in summary_2025.index else 0
    ext_2025 = int(summary_2025.loc[loc, 'External']) if loc in summary_2025.index else 0
    total_2025 = int_2025 + ext_2025
    
    int_2026 = int(summary_2026.loc[loc, 'Internal']) if loc in summary_2026.index else 0
    ext_2026 = int(summary_2026.loc[loc, 'External']) if loc in summary_2026.index else 0
    total_2026 = int_2026 + ext_2026
    
    change_abs = total_2026 - total_2025
    change_pct = (change_abs / total_2025 * 100) if total_2025 > 0 else 0
    
    summary_data.append({
        'Site': loc,
        'Internal_2025': int_2025,
        'External_2025': ext_2025,
        'Total_2025': total_2025,
        'Internal_2026': int_2026,
        'External_2026': ext_2026,
        'Total_2026': total_2026,
        'Change_Abs': change_abs,
        'Change_Pct': round(change_pct, 1)
    })

summary_df = pd.DataFrame(summary_data)

columns_order = [
    'Site',
    'Internal_2025', 'External_2025', 'Total_2025',
    'Internal_2026', 'External_2026', 'Total_2026',
    'Change_Abs', 'Change_Pct'
]
summary_df = summary_df[columns_order]

print("\n" + "=" * 100)
print("📊 AMP HC 汇总比较表 (2025 Jan vs 2026 Feb)")
print("=" * 100)
print(summary_df.to_string(index=False))

# 读取原始 Excel 文件
wb = load_workbook(file_path)

# 删除旧的 HC_Summary Sheet
if 'HC_Summary' in wb.sheetnames:
    del wb['HC_Summary']

# 创建新的汇总 Sheet
ws = wb.create_sheet('HC_Summary')

# 写入表头
headers = ['Site', 'Internal 2025', 'External 2025', 'Total 2025',
           'Internal 2026', 'External 2026', 'Total 2026',
           'Change (Abs)', 'Change (%)']

for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 写入数据
for row_idx, row in enumerate(summary_df.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 保存文件
wb.save(file_path)

print(f"\n✅ 汇总 Sheet 已更新：{file_path}")
print(f"📄 Sheet 名称：HC_Summary")
print(f"📊 共 {len(summary_df)} 个 Site")

# 保存 JSON 供地图使用（只显示 Total）
import json
map_data = {}
for _, row in summary_df.iterrows():
    site = row['Site']
    map_data[site] = {
        'Total_2025': int(row['Total_2025']),
        'Total_2026': int(row['Total_2026']),
    }

json_path = '/Users/liwang/.openclaw/workspace/Temp/amp-hc-data-v4.json'
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(map_data, f, indent=2, ensure_ascii=False)
print(f"💾 地图数据 JSON: {json_path}")

# 保存 CSV
csv_path = '/Users/liwang/.openclaw/workspace/Temp/amp-hc-summary-final-v2.csv'
summary_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
print(f"💾 CSV 备份：{csv_path}")
