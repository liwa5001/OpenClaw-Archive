#!/usr/bin/env python3
"""
修复 NaN Location 问题
将 NaN 行合并为"Unknown" Site
"""

import pandas as pd
import glob
import os
from openpyxl import load_workbook

# 文件目录
data_dir = '/Users/liwang/projects/CA_AMP_Workshop_Report'
os.chdir(data_dir)

# 查找文件
files = [f for f in glob.glob("*AMP*HC*.xlsx") if not f.startswith('~$')]
file_path = os.path.abspath(files[0])

# 读取原始数据
df_2025 = pd.read_excel(file_path, sheet_name='2025 Jan')
df_2026 = pd.read_excel(file_path, sheet_name='2026 Feb')

# 清理数据 - 填充 NaN 为'Unknown'
df_2025['Location'] = df_2025['Location'].fillna('Unknown')
df_2026['Location'] = df_2026['Location'].fillna('Unknown')
df_2025['Int/Ext'] = df_2025['Int/Ext'].fillna('Unknown')
df_2026['Int/Ext'] = df_2026['Int/Ext'].fillna('Unknown')

# 地点名称统一
df_2025.loc[df_2025['Location'] == 'Juarez', 'Location'] = 'Queretaro'
df_2026.loc[df_2026['Location'] == 'Juarez', 'Location'] = 'Queretaro'
df_2025.loc[df_2025['Location'] == 'Farmington Hills', 'Location'] = 'Novi'
df_2026.loc[df_2026['Location'] == 'Farmington Hills', 'Location'] = 'Novi'
df_2025.loc[df_2025['Location'] == 'Karlsbad', 'Location'] = 'Garching'
df_2026.loc[df_2026['Location'] == 'Karlsbad', 'Location'] = 'Garching'
df_2025.loc[df_2025['Location'] == 'Shanghai', 'Location'] = 'Suzhou'
df_2026.loc[df_2026['Location'] == 'Shanghai', 'Location'] = 'Suzhou'
df_2025.loc[df_2025['Location'] == 0, 'Location'] = 'Lodz'
df_2026.loc[df_2026['Location'] == 0, 'Location'] = 'Lodz'

print("📍 已处理：Juarez→Queretaro, Farmington Hills→Novi, Karlsbad→Garching, Shanghai→Suzhou, 0→Lodz, NaN→Unknown")

# 汇总 2025 Jan
summary_2025 = df_2025.groupby(['Location', 'Int/Ext'])['HC'].sum().unstack(fill_value=0)
if 'Internal' not in summary_2025.columns:
    summary_2025['Internal'] = 0
if 'External' not in summary_2025.columns:
    summary_2025['External'] = 0
summary_2025['Total_2025'] = summary_2025['Internal'] + summary_2025['External']

# 汇总 2026 Feb
summary_2026 = df_2026.groupby(['Location', 'Int/Ext'])['HC'].sum().unstack(fill_value=0)
if 'Internal' not in summary_2026.columns:
    summary_2026['Internal'] = 0
if 'External' not in summary_2026.columns:
    summary_2026['External'] = 0
summary_2026['Total_2026'] = summary_2026['Internal'] + summary_2026['External']

# 合并两年数据
all_locations = sorted(set(str(x) for x in summary_2025.index) | set(str(x) for x in summary_2026.index))

# 构建汇总 DataFrame
summary_data = []
for loc in all_locations:
    if loc == 'nan' or loc == '0':
        continue
    
    int_2025 = int(summary_2025.loc[loc, 'Internal']) if loc in summary_2025.index else 0
    ext_2025 = int(summary_2025.loc[loc, 'External']) if loc in summary_2025.index else 0
    total_2025 = int_2025 + ext_2025
    
    int_2026 = int(summary_2026.loc[loc, 'Internal']) if loc in summary_2026.index else 0
    ext_2026 = int(summary_2026.loc[loc, 'External']) if loc in summary_2026.index else 0
    total_2026 = int_2026 + ext_2026
    
    change_abs = total_2026 - total_2025
    change_pct = (change_abs / total_2025 * 100) if total_2025 > 0 else 0
    
    summary_data.append({
        'Site': loc,
        'Internal_2025': int_2025,
        'External_2025': ext_2025,
        'Total_2025': total_2025,
        'Internal_2026': int_2026,
        'External_2026': ext_2026,
        'Total_2026': total_2026,
        'Change_Abs': change_abs,
        'Change_Pct': round(change_pct, 1)
    })

summary_df = pd.DataFrame(summary_data)

columns_order = [
    'Site',
    'Internal_2025', 'External_2025', 'Total_2025',
    'Internal_2026', 'External_2026', 'Total_2026',
    'Change_Abs', 'Change_Pct'
]
summary_df = summary_df[columns_order]

print("\n" + "=" * 100)
print("📊 AMP HC 汇总比较表 (2025 Jan vs 2026 Feb)")
print("=" * 100)
print(summary_df.to_string(index=False))

# 验证总数
print("\n" + "=" * 60)
print("📊 总数验证")
print("=" * 60)
print(f"CSV 2025 Jan 总数：{summary_df['Total_2025'].sum()}")
print(f"CSV 2026 Feb 总数：{summary_df['Total_2026'].sum()}")
print(f"原始 2025 Jan 总数：{df_2025['HC'].sum()}")
print(f"原始 2026 Feb 总数：{df_2026['HC'].sum()}")

# 读取原始 Excel 文件
wb = load_workbook(file_path)

# 删除旧的 HC_Summary Sheet
if 'HC_Summary' in wb.sheetnames:
    del wb['HC_Summary']

# 创建新的汇总 Sheet
ws = wb.create_sheet('HC_Summary')

# 写入表头
headers = ['Site', 'Internal 2025', 'External 2025', 'Total 2025',
           'Internal 2026', 'External 2026', 'Total 2026',
           'Change (Abs)', 'Change (%)']

for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 写入数据
for row_idx, row in enumerate(summary_df.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 保存文件
wb.save(file_path)

print(f"\n✅ 汇总 Sheet 已更新：{file_path}")
print(f"📄 Sheet 名称：HC_Summary")
print(f"📊 共 {len(summary_df)} 个 Site")

# 保存 CSV
csv_path = '/Users/liwang/.openclaw/workspace/Temp/amp-hc-summary-fixed.csv'
summary_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
print(f"💾 CSV 备份：{csv_path}")
